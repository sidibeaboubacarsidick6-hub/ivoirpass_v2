"""
IvoirPass V2 — Administration du dashboard et wallet
"""
from django.contrib import admin
from django.contrib.admin.views.decorators import staff_member_required
from django.utils.html import format_html
from django.utils import timezone
from django.urls import path
from django.shortcuts import render
from django.db.models import Sum
from .models import OrganizerWallet, WalletTransaction, WithdrawalRequest, AuditLog, ReversalOTP, Dispute


# ============================================
# VUE RAPPORT BCEAO
# ============================================

@staff_member_required
def bceao_report_view(request):
    """Rapport mensuel BCEAO accessible dans l'admin."""
    from apps.tickets.models import Order
    from apps.store.models import ProductOrder
    from apps.accounts.models import CustomUser

    now = timezone.now()
    month_start = now.replace(day=1, hour=0, minute=0, second=0)

    ticket_orders = Order.objects.filter(paid_at__gte=month_start, status='paid').count()
    store_orders = ProductOrder.objects.filter(paid_at__gte=month_start, status='paid').count()
    ticket_volume = Order.objects.filter(paid_at__gte=month_start, status='paid').aggregate(t=Sum('total'))['t'] or 0
    store_volume = ProductOrder.objects.filter(paid_at__gte=month_start, status='paid').aggregate(t=Sum('total'))['t'] or 0
    withdrawals_count = WithdrawalRequest.objects.filter(created_at__gte=month_start).count()
    withdrawals_volume = WithdrawalRequest.objects.filter(created_at__gte=month_start, status='processed').aggregate(t=Sum('amount'))['t'] or 0
    total_users = CustomUser.objects.count()
    organizers = CustomUser.objects.filter(role='organizer').count()

    context = {
        'month': now.strftime('%B %Y'),
        'ticket_orders': ticket_orders, 'store_orders': store_orders,
        'total_transactions': ticket_orders + store_orders,
        'ticket_volume': int(ticket_volume), 'store_volume': int(store_volume),
        'total_volume': int(ticket_volume + store_volume),
        'withdrawals_count': withdrawals_count, 'withdrawals_volume': int(withdrawals_volume),
        'total_users': total_users, 'organizers': organizers,
    }
    return render(request, 'admin/bceao_report.html', context)


# ============================================
# INLINES
# ============================================

class WalletTransactionInline(admin.TabularInline):
    model = WalletTransaction
    extra = 0
    readonly_fields = ('type', 'amount', 'balance_after', 'description', 'reference', 'created_at')
    can_delete = False
    max_num = 0


# ============================================
# ORGANIZER WALLET
# ============================================

@admin.register(OrganizerWallet)
class OrganizerWalletAdmin(admin.ModelAdmin):
    list_display = ('organizer', 'balance_available', 'balance_pending', 'balance_withdrawn', 'preferred_payout_method', 'payout_phone')
    search_fields = ('organizer__email', 'organizer__first_name', 'payout_phone')
    readonly_fields = ('balance_withdrawn', 'created_at', 'updated_at')
    inlines = [WalletTransactionInline]


# ============================================
# WITHDRAWAL REQUEST
# ============================================

@admin.register(WithdrawalRequest)
class WithdrawalRequestAdmin(admin.ModelAdmin):
    list_display = ('reference', 'get_organizer', 'amount', 'payout_method', 'payout_phone', 'status_badge', 'created_at')
    list_filter = ('status', 'payout_method')
    search_fields = ('reference', 'wallet__organizer__email', 'payout_phone')
    readonly_fields = ('reference', 'amount_net', 'created_at', 'processed_at')
    actions = ['approve_requests', 'process_requests', 'reject_requests']

    def get_organizer(self, obj):
        return obj.wallet.organizer.get_full_name()
    get_organizer.short_description = "Organisateur"

    def status_badge(self, obj):
        colors = {'pending': '#F47920', 'approved': '#1B7A3E', 'processed': '#0dcaf0', 'rejected': '#dc3545'}
        color = colors.get(obj.status, '#6c757d')
        return format_html('<span style="background:{};color:white;padding:3px 10px;border-radius:20px;font-size:0.78rem;font-weight:700;">{}</span>', color, obj.get_status_display())
    status_badge.short_description = "Statut"

    @admin.action(description="✅ Approuver")
    def approve_requests(self, request, queryset):
        count = 0
        for wr in queryset.filter(status=WithdrawalRequest.Status.PENDING):
            wr.approve(admin_user=request.user, note="Approuvé via admin")
            count += 1
        self.message_user(request, f"{count} demande(s) approuvée(s).")

    @admin.action(description="💸 Marquer comme traitées")
    def process_requests(self, request, queryset):
        count = 0
        for wr in queryset.filter(status__in=[WithdrawalRequest.Status.PENDING, WithdrawalRequest.Status.APPROVED]):
            wr.mark_processed(admin_user=request.user, note="Virement effectué")
            count += 1
        self.message_user(request, f"{count} reversement(s) traité(s).")

    @admin.action(description="❌ Rejeter")
    def reject_requests(self, request, queryset):
        count = 0
        for wr in queryset.filter(status=WithdrawalRequest.Status.PENDING):
            wr.reject(admin_user=request.user, note="Rejeté via admin")
            count += 1
        self.message_user(request, f"{count} demande(s) rejetée(s).")


# ============================================
# DISPUTE
# ============================================

@admin.register(Dispute)
class DisputeAdmin(admin.ModelAdmin):
    list_display = ('reference', 'type', 'subject', 'reported_by', 'status_badge', 'created_at')
    list_filter = ('status', 'type')
    search_fields = ('reference', 'subject', 'reported_by__email', 'order_number')
    readonly_fields = ('reference', 'created_at', 'updated_at', 'resolved_at')

    def status_badge(self, obj):
        colors = {'open': '#F47920', 'investigating': '#0dcaf0', 'resolved': '#1B7A3E', 'closed': '#6c757d', 'rejected': '#dc3545'}
        color = colors.get(obj.status, '#6c757d')
        return format_html('<span style="background:{};color:white;padding:3px 10px;border-radius:20px;font-size:0.78rem;">{}</span>', color, obj.get_status_display())
    status_badge.short_description = "Statut"

    actions = ['mark_investigating', 'mark_resolved', 'mark_closed']

    @admin.action(description="🔍 En investigation")
    def mark_investigating(self, request, queryset):
        queryset.filter(status=Dispute.Status.OPEN).update(status=Dispute.Status.INVESTIGATING)
        self.message_user(request, "Litiges passés en investigation.")

    @admin.action(description="✅ Résoudre")
    def mark_resolved(self, request, queryset):
        queryset.exclude(status=Dispute.Status.CLOSED).update(status=Dispute.Status.RESOLVED, resolved_by=request.user, resolved_at=timezone.now())
        self.message_user(request, "Litiges résolus.")

    @admin.action(description="🔒 Fermer")
    def mark_closed(self, request, queryset):
        queryset.update(status=Dispute.Status.CLOSED)
        self.message_user(request, "Litiges fermés.")


# ============================================
# AUDIT LOG
# ============================================

@admin.register(AuditLog)
class AuditLogAdmin(admin.ModelAdmin):
    list_display = ('created_at', 'user', 'action_badge', 'model_name', 'description_truncated', 'ip_address')
    list_filter = ('action', 'model_name', 'user')
    search_fields = ('user__email', 'description', 'ip_address')
    readonly_fields = ('user', 'action', 'model_name', 'object_id', 'description', 'ip_address', 'created_at')
    date_hierarchy = 'created_at'
    ordering = ('-created_at',)
    list_per_page = 50

    def action_badge(self, obj):
        colors = {'create': '#1B7A3E', 'update': '#0dcaf0', 'delete': '#dc3545', 'publish': '#1B7A3E', 'unpublish': '#6c757d', 'login': '#6c757d', 'logout': '#6c757d', 'payout': '#F47920', 'export': '#6c757d', 'scan': '#1B7A3E', 'other': '#6c757d'}
        color = colors.get(obj.action, '#6c757d')
        return format_html('<span style="background:{};color:white;padding:2px 8px;border-radius:12px;font-size:0.78rem;">{}</span>', color, obj.get_action_display())
    action_badge.short_description = "Action"

    def description_truncated(self, obj):
        return obj.description[:100] + '...' if len(obj.description) > 100 else obj.description
    description_truncated.short_description = "Description"

    def has_add_permission(self, request):
        return False

    def has_change_permission(self, request, obj=None):
        return False

# ============================================
# EXPORTS BACK-OFFICE
# ============================================

from django.http import HttpResponse
import csv
import openpyxl
from io import BytesIO


@staff_member_required
def export_admin_csv(request):
    """Export CSV de toutes les données pour l'admin."""
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="ivoirpass_export_complet.csv"'
    response.write('\ufeff')
    writer = csv.writer(response)

    # Transactions
    writer.writerow(['=== TRANSACTIONS ==='])
    writer.writerow(['Type', 'Référence', 'Date', 'Montant', 'Statut'])

    from apps.tickets.models import Order
    for o in Order.objects.select_related('buyer').order_by('-created_at'):
        writer.writerow(['Billet', o.order_number, o.created_at.strftime('%d/%m/%Y'), int(o.total), o.get_status_display()])

    from apps.store.models import ProductOrder
    for o in ProductOrder.objects.select_related('buyer', 'product').order_by('-created_at'):
        writer.writerow(['Boutique', o.order_number, o.created_at.strftime('%d/%m/%Y'), int(o.total), o.get_status_display()])

    # Reversements
    writer.writerow([])
    writer.writerow(['=== REVERSEMENTS ==='])
    writer.writerow(['Référence', 'Organisateur', 'Montant', 'Méthode', 'Statut', 'Date'])
    for w in WithdrawalRequest.objects.select_related('wallet__organizer').order_by('-created_at'):
        writer.writerow([w.reference, w.wallet.organizer.get_full_name(), int(w.amount), w.get_payout_method_display(), w.get_status_display(), w.created_at.strftime('%d/%m/%Y')])

    # Utilisateurs
    writer.writerow([])
    writer.writerow(['=== UTILISATEURS ==='])
    writer.writerow(['Email', 'Nom', 'Rôle', 'Ville', 'Vérifié', 'Date inscription'])
    from apps.accounts.models import CustomUser
    for u in CustomUser.objects.order_by('-date_joined'):
        writer.writerow([u.email, u.get_full_name(), u.get_role_display(), u.city, 'Oui' if u.is_organizer_verified else 'Non', u.date_joined.strftime('%d/%m/%Y')])

    return response


@staff_member_required
def export_admin_excel(request):
    """Export Excel de toutes les données."""
    wb = openpyxl.Workbook()

    # Onglet Transactions
    ws1 = wb.active
    ws1.title = "Transactions"
    ws1.append(['Type', 'Référence', 'Date', 'Montant', 'Statut'])
    from apps.tickets.models import Order
    for o in Order.objects.select_related('buyer').order_by('-created_at')[:1000]:
        ws1.append(['Billet', o.order_number, o.created_at.strftime('%d/%m/%Y'), int(o.total), o.get_status_display()])
    from apps.store.models import ProductOrder
    for o in ProductOrder.objects.select_related('buyer', 'product').order_by('-created_at')[:1000]:
        ws1.append(['Boutique', o.order_number, o.created_at.strftime('%d/%m/%Y'), int(o.total), o.get_status_display()])

    # Onglet Reversements
    ws2 = wb.create_sheet("Reversements")
    ws2.append(['Référence', 'Organisateur', 'Montant', 'Méthode', 'Statut', 'Date'])
    for w in WithdrawalRequest.objects.select_related('wallet__organizer').order_by('-created_at'):
        ws2.append([w.reference, w.wallet.organizer.get_full_name(), int(w.amount), w.get_payout_method_display(), w.get_status_display(), w.created_at.strftime('%d/%m/%Y')])

    # Onglet Utilisateurs
    ws3 = wb.create_sheet("Utilisateurs")
    ws3.append(['Email', 'Nom', 'Rôle', 'Ville', 'Vérifié', 'Date inscription'])
    from apps.accounts.models import CustomUser
    for u in CustomUser.objects.order_by('-date_joined'):
        ws3.append([u.email, u.get_full_name(), u.get_role_display(), u.city, 'Oui' if u.is_organizer_verified else 'Non', u.date_joined.strftime('%d/%m/%Y')])

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="ivoirpass_export_complet.xlsx"'
    wb.save(response)
    return response